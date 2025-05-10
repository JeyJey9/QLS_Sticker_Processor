import os
import sys
import threading
import json
import pdfplumber
import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from tkinter import ttk, font
from collections import OrderedDict

# === SETTINGS FILE ===
SETTINGS_FILE = os.path.join(os.path.expanduser('~'), '.qls_sticker_settings.json')

# === LOAD SAVED PATHS ===
def load_settings():
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}

# === SAVE SETTINGS ===
def save_settings(paths_to_save):
    try:
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(paths_to_save, f, indent=2)
    except Exception as e:
        print(f"Could not save settings: {e}")

# === CONFIGURATION PLACEHOLDERS ===
mapping_file = None
master_files = {"BX726": None, "V769": None}
sheet_targets = {
    "BX726": {"CAL": "BX726 Plant Cont. Items_CAL 1", "WO CAL": "BX726 Plant Cont. Items_WO CAL1"},
    "V769": {"CAL": "V769 Plant Cont. Items_CAL", "WO CAL": "V769 Plant Cont. Items_WO CAL"}
}

# === CORE FUNCTIONS ===
def convert_pdf_to_excel(pdf_path, excel_path):
    try:
        tables = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                for tbl in page.extract_tables():
                    if tbl:
                        tables.extend(tbl)
        if tables:
            df = pd.DataFrame(tables)
            df.to_excel(excel_path, index=False, header=False)
            return True
    except Exception as e:
        print(f"Convert error {pdf_path}: {e}")
    return False


def extract_stickers_from_excel(path):
    stickers = OrderedDict()
    try:
        df = pd.read_excel(path, usecols=[0], header=None, skiprows=8, nrows=100)
        for v in df[0].dropna():
            line = str(v).split('\n')[0].strip()
            if any(x in line.upper() for x in ["LEFT","RIGHT","REAR","FRONT","PLANT","DATA","SYSTEM","CONFIDENTIAL","BUY OFF"]):
                continue
            if any(c.isdigit() for c in line):
                continue
            if len(line) < 5 or len(line.split()) > 6:
                continue
            stickers[line] = None
    except Exception as e:
        print(f"Extract error {path}: {e}")
    return list(stickers.keys())


def save_section_to_excel(section, out_path, results):
    any_data = False
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for prog in ["BX726", "V769"]:
            data = results[section][prog]
            if data:
                any_data = True
                max_len = max(len(r) for r in data)
                cols = ["PDF_File"] + [f"Sticker{i}" for i in range(1, max_len)]
                pd.DataFrame(data, columns=cols).to_excel(writer, sheet_name=prog, index=False)
    if not any_data and os.path.exists(out_path):
        os.remove(out_path)


def apply_mapping(master_path, map_path, results, out_folder):
    try:
        maps = pd.read_excel(map_path, sheet_name=None)
    except Exception as e:
        return f"Mapping load error: {e}"
    wb = openpyxl.load_workbook(master_path)
    base = os.path.basename(master_path)
    for t in ["CAL", "WO CAL"]:
        key = f"{os.path.splitext(base)[0]} {t}"
        if key not in maps:
            continue
        dfm = maps[key]
        program_key = os.path.splitext(base)[0]
        if program_key not in sheet_targets:
            continue
        sheet = sheet_targets[program_key][t]
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for _, row in dfm.iterrows():
            lbl = str(row.get("Manual_label")).strip().upper()
            if not lbl or lbl == 'NAN':
                continue
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(r, 2).value
                if cell and str(cell).strip().upper() == lbl:
                    for i, col in enumerate(range(23, 41)):
                        val = row.get(i + 2)
                        if pd.notna(val):
                            ws.cell(r, col, str(val))
                    break
    out_file = os.path.join(out_folder, f"[UPDATED] {base}")
    wb.save(out_file)
    return out_file


def process_all(base, map_file, out_folder, prog_var, log_widget):
    os.makedirs(out_folder, exist_ok=True)
    tmp = os.path.join(out_folder, "temp")
    os.makedirs(tmp, exist_ok=True)
    pdfs = [(r, f) for r, _, fs in os.walk(base) for f in fs if f.lower().endswith('.pdf')]
    total = len(pdfs)
    results = {"CAL": {"BX726": [], "V769": []}, "WO CAL": {"BX726": [], "V769": []}}
    for idx, (r, f) in enumerate(pdfs, 1):
        prog_var.set(int(idx / total * 100))
        path = os.path.join(r, f)
        xlsx = os.path.join(tmp, f[:-4] + '.xlsx')
        if not convert_pdf_to_excel(path, xlsx):
            continue
        sts = extract_stickers_from_excel(xlsx)
        if not sts:
            continue
        sec = "WO CAL" if 'wo cal' in r.lower() else 'CAL'
        prog = "V769" if 'v769' in r.lower() else 'BX726'
        results[sec][prog].append([f] + sts)
    # save
    save_section_to_excel("CAL", os.path.join(out_folder, "Extracted_Stickers_CAL.xlsx"), results)
    save_section_to_excel("WO CAL", os.path.join(out_folder, "Extracted_Stickers_WO_CAL.xlsx"), results)
    # apply
    for k, mp in master_files.items():
        if not mp:
            continue
        res = apply_mapping(mp, map_file, results, out_folder)
        log_widget.insert(tk.END, res + "\n")
    log_widget.insert(tk.END, "Done.\n")

# === GUI ===
root = tk.Tk()
root.title('QLS Sticker Processor')
root.geometry('900x650')  # Set window size

# Set global font
default_font = font.nametofont("TkDefaultFont")
default_font.config(size=12)
label_font = (default_font.actual('family'), 12)
entry_font = (default_font.actual('family'), 12)
button_font = (default_font.actual('family'), 12)
text_font = (default_font.actual('family'), 11)

# Load existing settings
settings = load_settings()

# File/folder browse functions
def browse_dir(entry, key):
    d = filedialog.askdirectory()
    if d:
        entry.delete(0, tk.END)
        entry.insert(0, d)
        save_vars[key].set(True)


def browse_file(entry, key):
    f = filedialog.askopenfilename(filetypes=[('Excel', '*.xlsx;*.xlsm;*.xls')])
    if f:
        entry.delete(0, tk.END)
        entry.insert(0, f)
        save_vars[key].set(True)

# Entries with save checkboxes
labels = ['PDF Base Folder', 'Sticker Mapping', 'Master BX726', 'Master V769', 'Output Folder']
keys = ['base', 'mapping', 'm1', 'm2', 'out']
rows = []
save_vars = {}
for i, (lbl, key) in enumerate(zip(labels, keys)):
    tk.Label(root, text=lbl + ':', font=label_font).grid(row=i, column=0, sticky='e', padx=10, pady=6)
    ent = tk.Entry(root, width=50, font=entry_font)
    ent.grid(row=i, column=1, padx=10, pady=6)
    if key in settings:
        ent.insert(0, settings[key])
    var = tk.BooleanVar(value=(key in settings))
    save_vars[key] = var
    cb = tk.Checkbutton(root, text='Save', variable=var, font=button_font)
    cb.grid(row=i, column=2, padx=10)
    mode = 'dir' if key in ['base', 'out'] else 'file'
    btn = tk.Button(root, text='Browse', font=button_font,
                    command=lambda e=ent, k=key, m=mode: browse_dir(e, k) if m == 'dir' else browse_file(e, k))
    btn.grid(row=i, column=3, padx=10)
    rows.append(ent)

# Progress bar & text log
prog = tk.DoubleVar()
pb = ttk.Progressbar(root, variable=prog, maximum=100, length=780)
pb.grid(row=6, column=0, columnspan=4, sticky='we', padx=10, pady=10)
log = scrolledtext.ScrolledText(root, width=106, height=12, font=text_font)
log.grid(row=7, column=0, columnspan=4, padx=10, pady=10)

# Run & Exit buttons
btn_run = tk.Button(root, text='Run', font=button_font, command=lambda: on_run())
btn_run.grid(row=8, column=2, pady=10)
exit_btn = tk.Button(root, text='Exit', font=button_font, command=root.destroy)
exit_btn.grid(row=8, column=3)

root.mainloop()
