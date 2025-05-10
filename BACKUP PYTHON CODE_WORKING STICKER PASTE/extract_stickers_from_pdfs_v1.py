
import os
import sys
import io
import pdfplumber
import pandas as pd
import openpyxl
import warnings
from collections import OrderedDict

# Suppress PDF noise
warnings.filterwarnings("ignore", category=UserWarning)
sys.stderr = io.StringIO()

# === CONFIGURATION ===
base_path = r"C:\Users\ghost\Documents\Project QLS\QLS\PDF"
converted_folder = os.path.join(base_path, "ConvertedExcels")
extracted_folder = os.path.join(base_path, "Extracted Files")
mapping_file = os.path.join(base_path, "Sticker_Mapping.xlsx")

master_files = {
    "BX726": os.path.join(base_path, "Study_NL_BX726 Plant controllable claims more than 1 in MY 25.xlsx"),
    "V769": os.path.join(base_path, "Study_NL_V769 Plant controllable claims more than 1 in MY 25.xlsx")
}

sheet_targets = {
    "BX726": {
        "CAL": "BX726 Plant Cont. Items_CAL 1",
        "WO CAL": "BX726 Plant Cont. Items_WO CAL1"
    },
    "V769": {
        "CAL": "V769 Plant Cont. Items_CAL",
        "WO CAL": "V769 Plant Cont. Items_WO CAL"
    }
}

os.makedirs(converted_folder, exist_ok=True)
os.makedirs(extracted_folder, exist_ok=True)

# Output paths for summaries
output_cal = os.path.join(extracted_folder, "Extracted_Stickers_CAL.xlsx")
output_wo_cal = os.path.join(extracted_folder, "Extracted_Stickers_WO_CAL.xlsx")

results = {
    "CAL": {"BX726": [], "V769": []},
    "WO CAL": {"BX726": [], "V769": []}
}

def convert_pdf_to_excel(pdf_path, excel_path):
    try:
        all_tables = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if table:
                        all_tables.extend(table)
        if all_tables:
            df = pd.DataFrame(all_tables)
            df.to_excel(excel_path, index=False, header=False)
            return True
    except:
        return False
    return False

def extract_stickers_from_excel(file_path):
    stickers = OrderedDict()
    try:
        df = pd.read_excel(file_path, usecols=[0], header=None, skiprows=8, nrows=100)
        for val in df[0].dropna():
            text = str(val).strip()
            first_line = text.split('\n')[0].strip()
            if any(x in first_line.upper() for x in ["LEFT", "RIGHT", "REAR", "FRONT", "PLANT", "DATA", "SYSTEM", "CONFIDENTIAL", "BUY OFF"]):
                continue
            if any(char.isdigit() for char in first_line):
                continue
            if len(first_line.split()) > 6 or len(first_line) < 5:
                continue
            stickers[first_line] = None
    except:
        pass
    return list(stickers.keys())

def print_progress(current, total, bar_length=30):
    percent = current / total
    filled = int(bar_length * percent)
    bar = '#' * filled + '-' * (bar_length - filled)
    print(f"\r[{bar}]  {current}/{total} files processed", end='')

# === Step 1: PDF Extraction ===
pdf_files = []
for root, _, files in os.walk(base_path):
    for file in files:
        if file.lower().endswith(".pdf"):
            pdf_files.append((root, file))

total_files = len(pdf_files)
print(f"Total PDFs to process: {total_files}\n")

processed = 0
for root, file in pdf_files:
    processed += 1
    print_progress(processed, total_files)

    pdf_path = os.path.join(root, file)
    excel_name = os.path.splitext(file)[0] + ".xlsx"
    excel_path = os.path.join(converted_folder, excel_name)

    converted = convert_pdf_to_excel(pdf_path, excel_path)
    if not converted:
        continue

    stickers = extract_stickers_from_excel(excel_path)
    if not stickers:
        os.remove(excel_path)
        continue

    entry = [file] + stickers

    section = "WO CAL" if "wo cal" in root.lower() else "CAL" if "cal" in root.lower() else None
    program = "V769" if "v769" in root.lower() else "BX726" if "bx726" in root.lower() else None

    if section and program:
        results[section][program].append(entry)

# === Step 2: Save Extracted Stickers ===
def save_section_to_excel(section_name, output_path):
    any_data = False
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for program in ["BX726", "V769"]:
            data = results[section_name][program]
            if data:
                any_data = True
                max_len = max(len(r) for r in data)
                columns = ["PDF_File"] + [f"Sticker{i}" for i in range(1, max_len)]
                df = pd.DataFrame(data, columns=columns)
                df.to_excel(writer, sheet_name=program, index=False)
                print(f"  {section_name} - {program}: {len(data)} rows saved.")
    if not any_data:
        try:
            os.remove(output_path)
        except:
            pass

print("\n\nSaving extracted sticker files...")
save_section_to_excel("CAL", output_cal)
save_section_to_excel("WO CAL", output_wo_cal)

# === Step 3: Apply Mapping ===
print("\nApplying sticker mappings to master files...")

sticker_col_start = 23
sticker_col_end = 40
mapping_sheets = pd.read_excel(mapping_file, sheet_name=None)

for program in ["BX726", "V769"]:
    if program not in master_files:
        continue
    master_path = master_files[program]
    wb = openpyxl.load_workbook(master_path)

    for cal_type in ["CAL", "WO CAL"]:
        sheet_name = sheet_targets[program][cal_type]
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found in {program}.")
            continue

        ws = wb[sheet_name]
        mapping_key = f"{program} {cal_type}"
        if mapping_key not in mapping_sheets:
            print(f"Mapping sheet '{mapping_key}' not found.")
            continue

        df = mapping_sheets[mapping_key]

        for _, row in df.iterrows():
            label = str(row.get("Manual_label")).strip().upper()
            if not label or label == "NAN":
                continue

            for r in range(2, ws.max_row + 1):
                master_label = str(ws.cell(r, 2).value).strip().upper() if ws.cell(r, 2).value else ""
                if master_label == label:
                    for i, col in enumerate(range(sticker_col_start, sticker_col_end + 1)):
                        val = row.get(i + 2)
                        if pd.notna(val):
                            ws.cell(r, col, str(val))
                    break

    updated_path = os.path.join(base_path, f"[UPDATED] {os.path.basename(master_path)}")
    wb.save(updated_path)
    print(f"  Updated file saved to: {updated_path}")

print("\nSummary:")
for section in ["CAL", "WO CAL"]:
    for program in ["BX726", "V769"]:
        count = len(results[section][program])
        print(f"  {section} - {program}: {count} PDF(s) processed")
